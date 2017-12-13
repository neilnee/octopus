# coding=utf-8
import calendar
import csv
import hashlib
import httplib
import json
import os
import time
from multiprocessing import Process

import datetime
from openpyxl import load_workbook
from scrapy import cmdline


class SpiderProcess(Process):
    def __init__(self, cmd):
        Process.__init__(self)
        self.cmd = cmd

    def run(self):
        print 'main spider process execute'
        cmdline.execute(self.cmd)


class Place4Week:
    # 场地信息
    p_name = ''
    p_code = ''
    wmq_name = ''
    wmq_code = ''
    ye_name = ''
    ye_code = ''
    ye_id = ''
    open_time = ''
    catchme_time = ''
    ret = 0
    cpt = 0

    # 测试数据
    test_gift = 0
    test_income = 0.0
    test_cost = 0.0
    test_play = 0

    # 周运营天数
    r_work_day = 0
    # 营收
    r_income = 0.0
    # 去娃娃盈利
    r_earn = 0.0
    # 抓取概率
    r_probability = 0.0
    # 掉落
    r_gift = 0
    # 游戏币
    r_coin_buy = 0
    # 派币
    r_coin_free = 0
    # 游戏次数
    r_play_time = 0
    # 渠道
    r_channel = ''
    # 扫码用户
    r_user_enter = 0
    # 充值用户
    r_user_pay = 0
    # 游戏用户
    r_user_play = 0
    # 扫码->充值转化率
    r_user_enter2pay = 0.0
    # 观影人次
    r_user_cinema = 0
    # 观影->扫码转化率
    r_user_cinema2enter = 0.0

    def __init__(self, place_info, mon, sun):
        self.p_name = place_info['name']
        self.p_code = md5(self.p_name)
        self.wmq_name = place_info['wmq_name']
        if len(self.wmq_name) > 0:
            self.wmq_code = md5(self.wmq_name)
        self.ye_name = place_info['ye_name']
        if len(self.ye_name) > 0:
            self.ye_code = md5(self.ye_name)
            self.ye_id = int(place_info['ye_id'])
        self.open_time = place_info['open_time']
        self.catchme_time = place_info['start']
        self.ret = int(place_info['ret'])
        self.test_gift = int(place_info['test_gift'])
        self.test_play = int(place_info['test_play'])
        self.test_income = float(place_info['test_income'])
        self.test_cost = float(place_info['test_cost'])

        if time.strptime(str(mon), '%Y-%m-%d') \
                <= time.strptime(self.catchme_time, '%Y-%m-%d') \
                <= time.strptime(str(sun), '%Y-%m-%d'):
            self.r_income = -self.test_income
            self.r_play_time = -self.test_play
            self.r_gift = -self.test_gift

    # noinspection PyBroadException
    def append_catchme(self, line, d):
        if time.strptime(self.open_time, '%Y-%m-%d') <= time.strptime(str(d), '%Y-%m-%d'):
            self.r_work_day += 1
            self.r_income += float(line[2].value)
            self.r_earn += float(line[13].value)
            self.r_gift += int(line[10].value)
            self.r_coin_buy += int(line[9].value) * self.cpt
            self.r_coin_free += 0
            self.r_play_time += int(line[9].value)
            self.r_user_enter += int(line[5].value)
            self.r_user_pay += int(line[7].value)
            self.r_user_play += int(line[6].value)
            try:
                self.r_probability = float(self.r_gift) / float(self.r_play_time)
                self.r_user_enter2pay = float(self.r_user_pay) / float(self.r_user_enter)
            except Exception:
                self.r_probability = 0.0
                self.r_user_enter2pay = 0.0

    # noinspection PyBroadException
    def append_wmq(self, line, d):
        if time.strptime(self.open_time, '%Y-%m-%d') <= time.strptime(str(d), '%Y-%m-%d'):
            self.r_income += float(line[1])
            self.r_earn += float(line[3])
            self.r_gift += int(line[6])
            self.r_coin_buy += int(line[7])
            self.r_coin_free += int(line[8])
            self.r_play_time += int(line[9])
            try:
                self.r_probability = float(self.r_gift) / float(self.r_play_time)
            except Exception:
                self.r_probability = 0.0
            pass

    # noinspection PyBroadException
    def append_cinema(self, cinema):
        self.r_user_cinema = cinema.audTotal
        try:
            self.r_user_cinema2enter = float(self.r_user_enter) / float(self.r_user_cinema)
        except Exception:
            self.r_user_cinema2enter = 0.0

    # noinspection PyBroadException
    def output(self):
        line = [0] * 18
        line[0] = str(self.p_name.encode('utf-8'))
        line[1] = self.r_work_day
        line[2] = self.r_income
        line[3] = self.r_earn
        line[4] = self.r_probability
        line[5] = self.r_gift
        line[6] = self.r_coin_buy
        line[7] = self.r_coin_free
        line[8] = self.r_play_time
        line[9] = str(self.r_channel.encode('utf-8'))
        line[10] = self.r_user_enter
        line[11] = self.r_user_pay
        line[12] = self.r_user_play
        line[13] = self.r_user_enter2pay
        line[14] = self.r_user_cinema
        line[15] = self.r_user_cinema2enter
        try:
            line[16] = float(self.r_play_time) / float(self.r_user_play)
        except Exception:
            line[16] = 0.0
        try:
            line[17] = float(self.r_income) / float(self.r_play_time)
        except Exception:
            line[17] = 0.0
        return line

    def print_infos(self):
        print('[' + self.r_channel.encode('utf-8') + '][' + self.p_name.encode('utf-8') + '][运营天数:'
              + str(self.r_work_day) + '][营收:' + str(self.r_income) + '][盈利:' + str(self.r_earn)
              + '][掉落:' + str(self.r_gift) + '][游戏币:' + str(self.r_coin_buy) + '][派币:' + str(self.r_coin_free)
              + '][游戏次数:' + str(self.r_play_time) + '][扫码用户:' + str(self.r_user_enter)
              + '][充值用户:' + str(self.r_user_pay) + '][游戏用户:' + str(self.r_user_play) + '][充值转化:'
              + str(self.r_user_enter2pay) + '][观影人次:' + str(self.r_user_cinema)
              + '][观影转化:' + str(self.r_user_cinema2enter) + ']')


class Channel4Week:
    ch_code = ''
    ch_key = ''
    places = {}
    cpt = 2
    ch_name = ''
    wmq_code_map = {}
    day_of_month = 0
    monday = ''
    sunday = ''

    def __init__(self, channel_info, mday, sday, dom):
        self.places = {}
        self.wmq_code_map = {}
        self.ch_code = channel_info['code']
        self.ch_key = md5(channel_info['name'])
        self.cpt = int(channel_info['cpt'])
        self.ch_name = channel_info['name']
        self.day_of_month = dom
        self.monday = mday
        self.sunday = sday
        if len(channel_info['include']) > 0:
            for p in channel_info['include']:
                if time.strptime(p['open_time'], '%Y-%m-%d') <= time.strptime(sday, '%Y-%m-%d'):
                    p4week = Place4Week(p, mday, sday)
                    p4week.cpt = self.cpt
                    p4week.r_channel = self.ch_name
                    self.places[p4week.p_code] = p4week
                    if len(p4week.wmq_code) > 0:
                        self.wmq_code_map[p4week.wmq_code] = p4week.p_code

    def append_catchme(self, line, d):
        p_code = md5(line[0].value, True)
        if p_code in self.places.keys():
            self.places[p_code].append_catchme(line, d)

    def append_weimaqi(self, line, d):
        wmq_code = md5(line[0], False)
        if wmq_code in self.wmq_code_map.keys():
            if self.wmq_code_map[wmq_code] in self.places.keys():
                self.places[self.wmq_code_map[wmq_code]].append_wmq(line, d)

    # noinspection PyBroadException
    def output(self):
        l_income = 0.0
        l_earn = 0.0
        l_probability = 0.0
        l_gift = 0
        l_coin_buy = 0
        l_coin_free = 0
        l_play_time = 0
        l_user_enter = 0
        l_user_pay = 0
        l_user_play = 0
        l_user_enter2pay = 0.0
        l_user_cinema = 0
        l_user_cinema2enter = 0.0
        l_ret = 0.0

        valid_user_enter = 0
        valid_user_cinema = 0

        for p in self.places.values():
            l_income += p.r_income
            l_earn += p.r_earn
            l_gift += p.r_gift
            l_coin_buy += p.r_coin_buy
            l_coin_free += p.r_coin_free
            l_play_time += p.r_play_time
            l_user_enter += p.r_user_enter
            l_user_pay += p.r_user_pay
            l_user_play += p.r_user_play
            l_ret += p.r_work_day * p.ret / self.day_of_month
            l_user_cinema += p.r_user_cinema

            if p.r_user_cinema > 0 \
                    and time.strptime(p.catchme_time, '%Y-%m-%d') <= time.strptime(str(sunday), '%Y-%m-%d'):
                valid_user_cinema += p.r_user_cinema
                valid_user_enter += p.r_user_enter
                try:
                    l_user_cinema2enter = float(valid_user_enter) / float(valid_user_cinema)
                except Exception:
                    pass
            try:
                l_user_enter2pay = float(l_user_pay) / float(l_user_enter)
            except Exception:
                pass
            try:
                l_probability = float(l_gift) / float(l_play_time)
            except Exception:
                pass

        line = [0] * 18
        line[0] = str(self.ch_name.encode('utf-8'))
        line[1] = '/'
        line[2] = l_income
        line[3] = l_earn
        line[4] = l_probability
        line[5] = l_gift
        line[6] = l_coin_buy
        line[7] = l_coin_free
        line[8] = l_play_time
        line[9] = '/'
        line[10] = l_user_enter
        line[11] = l_user_pay
        line[12] = l_user_play
        line[13] = l_user_enter2pay
        line[14] = l_user_cinema
        line[15] = l_user_cinema2enter
        try:
            line[16] = float(l_play_time) / float(l_user_play)
        except Exception:
            line[16] = 0.0
        try:
            line[17] = float(l_income) / float(l_play_time)
        except Exception:
            line[17] = 0.0
        return line, valid_user_enter, valid_user_cinema


class Cinema4Week:
    cinemaCode = ''
    cinemaId = ''  # 影院ID
    cinemaName = ''  # 影院名称
    amount = 0.0  # 当周票房
    scenes = 0.0  # 当周场次
    avgScreen = 0.0  # 单荧幕平均周票房
    avgPS = 0.0  # 场均人次
    screen_yield = 0.0  # 单日单厅票房
    scenes_time = 0.0  # 单日单厅场次
    audTotal = 0  # 周观影人次

    def __init__(self, line):
        self.cinemaId = int(line[0])
        self.cinemaName = line[1]
        self.amount = float(line[2])
        self.scenes = float(line[3])
        self.avgScreen = float(line[4])
        self.avgPS = float(line[5])
        self.screen_yield = float(line[6])
        self.scenes_time = float(line[7])
        self.audTotal = self.scenes * self.avgPS
        self.cinemaCode = md5(self.cinemaName, False)
        pass


class Channel4Day:
    ch_code = ''
    ch_key = ''
    places = []
    revenue = {}
    revenue_wmq = []
    cpt = 2
    revenue_cvt = {}
    ret = 0
    name = ''
    with_wmq = True
    close_place = []

    def __init__(self, channel_info, ysd):
        self.ch_code = channel_info['code']
        self.ch_key = md5(channel_info['name'])
        self.cpt = int(channel_info['cpt'])
        self.places = []
        self.revenue = {}
        self.revenue_wmq = []
        self.revenue_cvt = {}
        self.name = channel_info['name']
        struct_t = time.strptime(ysd, '%Y-%m-%d')
        self.with_wmq = bool(channel_info['wmq'])

        if len(channel_info['include']) > 0:
            for place in channel_info['include']:
                if time.strptime(ysd, '%Y-%m-%d') >= time.strptime(place['start'], '%Y-%m-%d'):
                    place_key = md5(place['name'])
                    wmq_key = md5(place['wmq_name'])
                    self.revenue_cvt[wmq_key] = place_key
                    if bool(place['open']):
                        self.places.append(place_key)
                        self.revenue[place_key] = [0] * 19
                        self.revenue[place_key][0] = str(place['name'].encode('utf-8'))
                        self.revenue[place_key][11] = int(place['disable'])
                        self.revenue[place_key][12] = str(self.name.encode('utf-8'))
                        if place['start'] == ysd:
                            self.revenue[place_key][1] = -float(place['test_income'])
                            self.revenue[place_key][3] = -float(place['test_income']) + float(place['test_cost'])
                            self.revenue[place_key][6] = -float(place['test_gift'])
                            self.revenue[place_key][9] = -float(place['test_play'])
                            self.revenue[place_key][7] = -int(place['test_play'] * self.cpt)
                    else:
                        self.close_place.append(place_key)
                if time.strptime(ysd, '%Y-%m-%d') >= time.strptime(place['open_time'], '%Y-%m-%d'):
                    self.ret += float(place['ret'])

        self.ret = self.ret / calendar.monthrange(struct_t.tm_year, struct_t.tm_mon)[1]

    # noinspection PyBroadException
    def append_weimaqi_data(self, input_line):
        wmq_key = md5(input_line[0], False)
        if wmq_key in self.revenue_cvt.keys() and self.revenue_cvt[wmq_key] in self.revenue.keys():
            revenue_line = self.revenue[self.revenue_cvt[wmq_key]]
            revenue_line[1] += float(input_line[1])
            revenue_line[3] += float(input_line[3])
            revenue_line[6] += int(input_line[6])
            revenue_line[9] += int(input_line[9])
            revenue_line[7] += int(input_line[7])
            revenue_line[8] += int(input_line[8])
            try:
                revenue_line[5] = float(revenue_line[6]) / float(revenue_line[9])
            except Exception:
                revenue_line[5] = 0.0
            if int(revenue_line[10]) == 0 and int(revenue_line[11]) == 0:
                revenue_line[10] = int(input_line[10])
                revenue_line[11] = int(input_line[11])
            revenue_line[2] = float(revenue_line[1]) / (revenue_line[10] + revenue_line[11])
            revenue_line[4] = revenue_line[3] / (revenue_line[10] + revenue_line[11])
            try:
                revenue_line[17] = float(revenue_line[9]) / float(revenue_line[15])
            except Exception:
                revenue_line[17] = 0.0
            try:
                revenue_line[18] = float(revenue_line[1]) / float(revenue_line[9])
            except Exception:
                revenue_line[18] = 0.0
        else:
            if wmq_key in self.revenue_cvt.keys() and self.revenue_cvt[wmq_key] in self.close_place:
                return
            wmq_line = [0] * 19
            wmq_line[0] = input_line[0]
            wmq_line[1] = input_line[1]
            wmq_line[2] = input_line[2]
            wmq_line[3] = input_line[3]
            wmq_line[4] = input_line[4]
            wmq_line[5] = input_line[5]
            wmq_line[6] = input_line[6]
            wmq_line[7] = input_line[7]
            wmq_line[8] = input_line[8]
            wmq_line[9] = input_line[9]
            wmq_line[10] = input_line[10]
            wmq_line[11] = input_line[11]
            wmq_line[12] = str(self.name.encode('utf-8'))
            wmq_line[13] = 0
            wmq_line[14] = 0
            wmq_line[15] = 0
            wmq_line[16] = 0
            wmq_line[17] = 0.0
            wmq_line[18] = 0.0
            self.revenue_wmq.append(wmq_line)

    # noinspection PyBroadException
    def append_catchme_data(self, input_line):
        place_key = md5(input_line[0].value, True)
        if place_key in self.revenue.keys():
            revenue_line = self.revenue[place_key]
            # 营收
            revenue_line[1] += float(input_line[2].value)
            # 盈利(去礼品)
            revenue_line[3] += float(input_line[13].value)
            # 掉落
            revenue_line[6] += int(input_line[10].value)
            # 游戏次数
            revenue_line[9] += int(input_line[9].value)
            # 游戏币
            revenue_line[7] += int(input_line[9].value) * self.cpt
            # 免费发放游戏币
            revenue_line[8] = 0
            try:
                # 抓取概率
                revenue_line[5] = float(revenue_line[6]) / float(revenue_line[9])
            except Exception:
                revenue_line[5] = 0.0
            # 在线台数
            revenue_line[10] += int(input_line[4].value)
            # 离线台数
            revenue_line[11] += int(input_line[3].value) - int(input_line[4].value)
            # 台均营收
            try:
                revenue_line[2] = float(revenue_line[1]) / (revenue_line[10] + revenue_line[11])
            except Exception:
                revenue_line[2] = 0.0
            # 台均盈利(去礼品)
            try:
                revenue_line[4] = revenue_line[3] / (revenue_line[10] + revenue_line[11])
            except Exception:
                revenue_line[4] = 0.0
            # 扫码用户
            revenue_line[13] += int(input_line[5].value)
            # 充值用户
            revenue_line[14] += int(input_line[7].value)
            # 游戏用户
            revenue_line[15] += int(input_line[6].value)
            # 扫码->充值转化率
            try:
                revenue_line[16] = float(revenue_line[14]) / float(revenue_line[13])
            except Exception:
                revenue_line[16] = 0.0
            # 人均游戏次数
            try:
                revenue_line[17] = float(revenue_line[9]) / float(revenue_line[15])
            except Exception:
                revenue_line[17] = 0.0
            # 单次游戏均价
            try:
                revenue_line[18] = float(revenue_line[1]) / float(revenue_line[9])
            except Exception:
                revenue_line[18] = 0.0
            print('[' + revenue_line[12] + '][' + revenue_line[0] + '][' + str(revenue_line[1])
                  + '][' + str(input_line[3].value) + '][' + str(input_line[4].value) + ']')

    # noinspection PyBroadException
    def calculat_total_revenue(self):
        income = 0.0
        income_ave = 0.0
        profit = 0.0
        profit_ave = 0.0
        profit_net = 0.0
        profit_net_ave = 0.0
        profit_net_percent = 0.0
        probability = 0.0
        total_device = 0
        total_place = 0
        total_coin_buy = 0
        total_coin_free = 0
        total_device_online = 0
        total_device_offline = 0
        total_352 = 0

        total_play = 0
        total_gift = 0

        total_user_enter = 0
        total_user_pay = 0
        total_user_play = 0
        user_enter_to_pay = 0.0

        total_line = []

        if len(self.revenue) > 0:
            for place in self.revenue.values():
                income += float(place[1])
                profit += float(place[3])
                total_device_online += int(place[10])
                total_device_offline += int(place[11])
                total_device += int(place[10]) + int(place[11])
                total_place += 1
                total_play += int(place[9])
                total_gift += int(place[6])
                total_coin_buy += int(place[7])
                total_coin_free += int(place[8])
                if (float(place[10]) + float(place[11])) > 0 and float(place[1]) / (float(place[10]) + float(place[11])) >= 35.2:
                    total_352 += 1
                total_user_enter += int(place[13])
                total_user_pay += int(place[14])
                total_user_play += int(place[15])
        if len(self.revenue_wmq) > 0:
            for place in self.revenue_wmq:
                income += float(place[1])
                profit += float(place[3])
                total_device_online += int(place[10])
                total_device_offline += int(place[11])
                total_device += int(place[10]) + int(place[11])
                total_place += 1
                total_play += int(place[9])
                total_gift += int(place[6])
                total_coin_buy += int(place[7])
                total_coin_free += int(place[8])
                if (float(place[10]) + float(place[11])) > 0 and float(place[1]) / (float(place[10]) + float(place[11])) >= 35.2:
                    total_352 += 1
        if total_device > 0:
            income_ave = income / total_device
            profit_ave = profit / total_device
            profit_net = profit - self.ret
            profit_net_ave = profit_net / total_device
            if income > 0:
                profit_net_percent = profit_net / income
            if total_play > 0:
                probability = float(total_gift) / float(total_play)
            try:
                user_enter_to_pay = float(total_user_pay) / float(total_user_enter)
            except Exception:
                user_enter_to_pay = 0.0

            total_line = [0] * 25
            total_line[0] = str(self.name.encode('utf-8'))
            total_line[1] = income
            total_line[2] = income_ave
            total_line[3] = profit
            total_line[4] = profit_ave
            total_line[5] = probability
            total_line[6] = total_gift
            total_line[7] = total_coin_buy
            total_line[8] = total_coin_free
            total_line[9] = total_play
            total_line[10] = total_device_online
            total_line[11] = total_device_offline
            total_line[12] = total_device_online + total_device_offline
            total_line[13] = total_user_enter
            total_line[14] = total_user_pay
            total_line[15] = total_user_play
            total_line[16] = user_enter_to_pay
            try:
                total_line[17] = float(total_play) / float(total_user_play)
            except Exception:
                total_line[17] = 0.0
            try:
                total_line[18] = float(income) / float(total_play)
            except Exception:
                total_line[18] = 0.0
            total_line[19] = self.ret
            total_line[20] = profit_net
            total_line[21] = profit_net_ave
            total_line[22] = total_place
            total_line[23] = total_352
            total_line[24] = float(total_352) / float(total_place)

        return self.name.encode('utf-8'), income, income_ave, profit, profit_ave, profit_net, profit_net_ave, \
               profit_net_percent, probability, total_device, total_place, total_play, total_gift, total_line, \
               total_user_enter, total_user_pay, total_user_play, user_enter_to_pay


# 加载catchme数据时需要包含的渠道场地
includes = {}


def md5(input_str, encode=True):
    if encode:
        return hashlib.md5(input_str.encode('utf-8').strip()).hexdigest()
    else:
        return hashlib.md5(input_str.strip()).hexdigest()


def getyesterday():
    today = datetime.date.today()
    oneday = datetime.timedelta(days=1)
    return today - oneday


def load_weimaqi(input_file, ch_item):
    with open(input_file) as wmq_file:
        idx = 0
        for line in wmq_file:
            idx = idx + 1
            if idx == 1:
                continue
            line = line.strip()
            line = line.replace('\r\n', '')
            line_list = line.split(',')
            ch_item.append_weimaqi_data(line_list)


class Cinema:
    cinemaId = ''  # 影院ID
    cinemaName = ''  # 影院名称
    amount = 0.0  # 当周票房
    scenes = 0.0  # 当周场次
    avgScreen = 0.0  # 单荧幕平均周票房
    avgPS = 0.0  # 场均人次
    screen_yield = 0.0  # 单日单厅票房
    scenes_time = 0.0  # 单日单厅场次

    def __init__(self, json_cinema):
        self.cinemaId = str(json_cinema['cinemaId'])
        self.cinemaName = str(json_cinema['cinemaName'].encode('utf-8'))
        self.amount = float(json_cinema['amount'])
        self.scenes = float(json_cinema['scenes'])
        self.avgScreen = float(json_cinema['avgScreen'])
        self.avgPS = float(json_cinema['avgPS'])
        self.screen_yield = float(json_cinema['screen_yield'])
        self.scenes_time = float(json_cinema['scenes_time'])

    def output(self):
        output = [0] * 8
        output[0] = self.cinemaId
        output[1] = self.cinemaName
        output[2] = self.amount
        output[3] = self.scenes
        output[4] = self.avgScreen
        output[5] = self.avgPS
        output[6] = self.screen_yield
        output[7] = self.scenes_time
        return output


def do_request_cinema_get(idx, week):
    ret = []
    conn = httplib.HTTPConnection("www.cbooo.cn")
    conn.request("GET", "/BoxOffice/getCBW?pIndex=" + str(idx) + "&dt=" + str(week))
    response = conn.getresponse()
    # print "request response : " + str(response.status)
    cinemas = json.loads(response.read())['data1']
    for i in range(0, len(cinemas)):
        ret.append(Cinema(cinemas[i]))
    conn.close()
    return ret


def request_cinema_data(w_idx):
    cinema_writer = csv.writer(file('revenue/cbooo/week_' + str(w_idx) + '.csv', 'wb'))
    title = [0] * 8
    title[0] = '影院ID'
    title[1] = '影院名称'
    title[2] = '当周票房'
    title[3] = '当周场次'
    title[4] = '单荧幕平均周票房'
    title[5] = '场均人次'
    title[6] = '单日单厅票房'
    title[7] = '单日单厅场次'
    cinema_writer.writerow(title)
    index = 1
    count = 0
    ids = []
    while index > 0:
        cinema_list = do_request_cinema_get(index, w_idx)
        for cinema in cinema_list:
            if cinema.cinemaId not in ids:
                ids.append(cinema.cinemaId)
                cinema_writer.writerow(cinema.output())
                count += 1
                print('影院[' + cinema.cinemaName + ']导入完成;总数[' + str(count) + ']')
        if len(cinema_list) >= 10:
            index += 1
        else:
            index = -1


if __name__ == '__main__':
    yesterday = str(getyesterday())

    with open("../channel.json") as ch_file:
        load_ch = json.load(ch_file)
        for ch in load_ch:
            if ch['name']:
                includes[md5(ch['name'])] = Channel4Day(ch, yesterday)

    cmds = []
    with open("../account.json") as account_file:
        load_account = json.load(account_file)
        for i in range(0, len(load_account)):
            wmq_ch_file = 'revenue/weimaqi/revenue_' + yesterday + '_' + load_account[i]['place'] + '.csv'
            if not os.path.isfile(wmq_ch_file):
                cmds.append(('scrapy crawl weimaqi -a uid='
                             + str(load_account[i]['uid'])
                             + ' -a pwd='
                             + str(load_account[i]['pwd'])
                             + ' -a yesterday='
                             + yesterday
                             + ' -a price=12'
                             + ' -a ch='
                             + str(load_account[i]['place'])
                             + ' -a cpt='
                             + str(load_account[i]['cpt'])
                             + ' -o '
                             + 'revenue/weimaqi/revenue' + '_' + yesterday + '_' + str(load_account[i]['place'])
                             + '.csv').split())
        account_file.close()
    for c in cmds:
        p = SpiderProcess(c)
        p.daemon = True
        p.start()
        p.join()

    catchme_file = 'revenue/catchme_ch/revenue_' + yesterday + '_ch.xlsx'
    if os.path.isfile(catchme_file):
        workbook = load_workbook(catchme_file)
        worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
        for row in worksheet.rows:
            if row[1].value:
                c_key = md5(row[1].value, True)
                if c_key in includes.keys():
                    p_key = md5(row[0].value, True)
                    if p_key in includes[c_key].places:
                        includes[c_key].append_catchme_data(row)
        workbook.close()

    t_income = 0.0
    t_income_ave = 0.0
    t_profit = 0.0
    t_profit_ave = 0.0
    t_profit_net = 0.0
    t_profit_net_ave = 0.0
    t_profit_net_percent = 0.0
    t_probability = 0.0
    t_total_device = 0
    t_total_place = 0
    t_total_play = 0
    t_total_gift = 0

    place_output = []
    channel_output_line = []

    writer = csv.writer(file('revenue/revenue_' + yesterday + '.csv', 'wb'))
    for item in includes.values():
        if item.with_wmq:
            load_weimaqi('revenue/weimaqi/revenue_' + yesterday + '_' + item.ch_code + '.csv',
                         includes[md5(item.name)])

        out_name, out_income, out_income_ave, out_profit, out_profit_ave, out_profit_net, out_profit_net_ave, \
        out_profit_net_percent, out_probability, out_total_device, out_total_place, out_play, out_gift, out_line, \
        out_user_enter, out_user_pay, out_user_play, out_user_enter2pay = item.calculat_total_revenue()

        if out_total_device == 0:
            continue

        if len(out_line) > 0:
            channel_output_line.append(out_line)

        writer.writerows(item.revenue.values())
        if item.with_wmq:
            writer.writerows(item.revenue_wmq)

        t_income += out_income
        t_profit += out_profit
        t_profit_net += out_profit_net
        t_total_device += out_total_device
        t_total_place += out_total_place
        t_total_play += out_play
        t_total_gift += out_gift
        t_income_ave = t_income / t_total_device
        t_profit_ave = t_profit / t_total_device
        t_profit_net_ave = t_profit_net / t_total_device
        t_profit_net_percent = t_profit_net / t_income
        t_probability = float(t_total_gift) / float(t_total_play)

        place_output.append('\n[渠道数据 - ' + out_name
                            + ']\n营收: ' + str(round(out_income, 1))
                            + '\n台均营收: ' + str(round(out_income_ave, 1))
                            + '\n盈利(去除娃娃和租金): ' + str(round(out_profit_net, 1))
                            + '\n台均盈利: ' + str(round(out_profit_net_ave, 1))
                            + '\n盈利率: ' + str(round(out_profit_net_percent * 100, 2))
                            + '%\n抓取概率: ' + str(round(out_probability * 100, 2))
                            + '%\n总台数: ' + str(out_total_device)
                            + '\n总场地数: ' + str(out_total_place))

    writer.writerow('')
    writer.writerows(channel_output_line)

    time.sleep(3)

    print ('\n[汇总数据]'
           + '\n营收: ' + str(round(t_income, 1))
           + '\n台均营收: ' + str(round(t_income_ave, 1))
           + '\n盈利(去除娃娃和租金): ' + str(round(t_profit_net, 1))
           + '\n台均盈利: ' + str(round(t_profit_net_ave, 1))
           + '\n盈利率: ' + str(round(t_profit_net_percent * 100, 2))
           + '%\n抓取概率: ' + str(round(t_probability * 100, 2))
           + '%\n总台数: ' + str(t_total_device)
           + '\n总场地数: ' + str(t_total_place))

    for output_line in place_output:
        print (output_line)

    # 开始计算周汇总数据
    week_chs = {}
    cinema_map = {}
    date978 = datetime.date(2017, 12, 03)

    if time.strptime(yesterday, '%Y-%m-%d').tm_wday == 6:
        sunday = yesterday
        struct_sunday = time.strptime(sunday, '%Y-%m-%d')
        datetime_sunday = datetime.date(struct_sunday.tm_year, struct_sunday.tm_mon, struct_sunday.tm_mday)
        gap_day = datetime_sunday - date978
        monday = str(datetime_sunday - datetime.timedelta(days=6))
        week_idx = 978 + gap_day.days / 7
        dayofm = calendar.monthrange(struct_sunday.tm_year, struct_sunday.tm_mon)[1]
        print('start week calculate: mon[' + monday + '] - sun [' + sunday + '] & week_idx [' + str(week_idx) + ']')

        with open("../channel.json") as ch_file:
            load_ch = json.load(ch_file)
            for ch in load_ch:
                if ch['name']:
                    week_chs[md5(ch['name'])] = Channel4Week(ch, monday, sunday, dayofm)

        for i in range(0, 7):
            day = datetime_sunday - datetime.timedelta(days=i)
            # 加载catchme数据
            catchme_day_file = 'revenue/catchme_ch/revenue_' + str(day) + '_ch.xlsx'
            if os.path.isfile(catchme_day_file):
                workbook = load_workbook(catchme_day_file)
                worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
                for row in worksheet.rows:
                    if row[1].value:
                        c_key = md5(row[1].value, True)
                        if c_key in week_chs.keys():
                            week_chs[c_key].append_catchme(row, day)
                workbook.close()
            # 加载wmq数据
            wmq_day_file_hd = 'revenue/weimaqi/revenue_' + str(day) + '_hd.csv'
            if os.path.isfile(wmq_day_file_hd):
                with open(wmq_day_file_hd) as hd_file:
                    ignore = True
                    for li in hd_file:
                        if ignore:
                            ignore = False
                            continue
                        li = li.strip()
                        li = li.replace('\r\n', '')
                        li_list = li.split(',')
                        week_chs[md5('恒大', False)].append_weimaqi(li_list, day)
            wmq_day_file_xh = 'revenue/weimaqi/revenue_' + str(day) + '_xh.csv'
            if os.path.isfile(wmq_day_file_xh):
                with open(wmq_day_file_xh) as xh_file:
                    ignore = True
                    for li in xh_file:
                        if ignore:
                            ignore = False
                            continue
                        li = li.strip()
                        li = li.replace('\r\n', '')
                        li_list = li.split(',')
                        week_chs[md5('星河', False)].append_weimaqi(li_list, day)

        cbooo_path = 'revenue/cbooo/week_' + str(week_idx) + '.csv'
        if not os.path.isfile(cbooo_path):
            request_cinema_data(week_idx)

        if os.path.isfile(cbooo_path):
            with open(cbooo_path) as cbooo_file:
                ignore = True
                for li in cbooo_file:
                    if ignore:
                        ignore = False
                        continue
                    li = li.strip()
                    li = li.replace('\r\n', '')
                    li_list = li.split(',')
                    c = Cinema4Week(li_list)
                    cinema_map[c.cinemaCode] = c

        final_writer = csv.writer(file('revenue/week/week_' + str(monday) + '_' + str(sunday) + '.csv', 'wb'))
        for channel in week_chs.values():
            for place in channel.places.values():
                if place.ye_code in cinema_map.keys():
                    place.append_cinema(cinema_map[place.ye_code])
                final_writer.writerow(place.output())
        final_writer.writerow('')
        total_valid_user_enter = 0
        total_valid_user_cinema = 0
        total_line = [0] * 18
        total_line[0] = '汇总'
        total_line[1] = '/'
        total_line[9] = '/'
        for channel in week_chs.values():
            line, valid_user_enter, valid_user_cinema = channel.output()
            total_valid_user_enter += valid_user_enter
            total_valid_user_cinema += valid_user_cinema
            total_line[2] += float(line[2])
            total_line[3] += float(line[3])
            total_line[5] += int(line[5])
            total_line[6] += int(line[6])
            total_line[7] += int(line[7])
            total_line[8] += int(line[8])
            total_line[10] += int(line[10])
            total_line[11] += int(line[11])
            total_line[12] += int(line[12])
            total_line[14] += float(line[14])
            final_writer.writerow(line)
        # noinspection PyBroadException
        try:
            total_line[4] = float(total_line[5]) / float(total_line[8])
        except Exception:
            total_line[4] = 0.0
        # noinspection PyBroadException
        try:
            total_line[13] = float(total_line[11]) / float(total_line[10])
        except Exception:
            total_line[13] = 0.0
        # noinspection PyBroadException
        try:
            total_line[15] = float(total_valid_user_enter) / float(total_valid_user_cinema)
        except Exception:
            total_line[15] = 0.0
        # noinspection PyBroadException
        try:
            total_line[16] = float(total_line[8]) / float(total_line[12])
        except Exception:
            total_line[16] = 0.0
        # noinspection PyBroadException
        try:
            total_line[17] = float(total_line[2]) / float(total_line[8])
        except Exception:
            total_line[17] = 0.0

        final_writer.writerow(total_line)
